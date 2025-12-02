import openpyxl

file = openpyxl.load_workbook("NewXLFile.xlsx")

file_data = file["main"]

p_list = {}

for i in range(2, file_data.max_row + 1):
    pass_data = file_data.cell(i, 3).value

    if pass_data in p_list:
        p_list[pass_data] = p_list[pass_data]+1

    else:
        p_list[pass_data] = 1

print(p_list)


